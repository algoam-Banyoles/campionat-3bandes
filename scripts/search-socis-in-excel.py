"""
Cerca els números de soci "fantasma" (fotos sense coincidència a la BD) dins
dels fitxers Excel històrics del Foment. Prioritza GestioSocials.xlsm com a
font principal.

Ús: python scripts/search-socis-in-excel.py
"""
from __future__ import annotations

import sys
from pathlib import Path
from openpyxl import load_workbook

HISTORIC = Path("G:/Mi unidad/Foment/Historic")
TARGETS = {1111, 8280, 8335, 8435, 8560, 8694}

# Prioritzem els Excel que probablement tenen la llista mestra de socis
PRIORITY_FILES = [
    HISTORIC / "GestioSocials.xlsm",
]


def scan_workbook(path: Path, targets: set[int]) -> list[tuple[int, str, str, int, str]]:
    """Retorna llista de (numero_trobat, sheet, cel·la, row, context_text)."""
    results: list[tuple[int, str, str, int, str]] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    except Exception as e:
        print(f"  ⚠️  No puc obrir {path.name}: {e}")
        return results

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                # Try to match as integer
                try:
                    n = int(v) if isinstance(v, (int, float)) else int(str(v).strip())
                except (ValueError, TypeError):
                    continue
                if n in targets:
                    # Capture row context for readability
                    row_values = [str(c.value).strip() if c.value is not None else "" for c in row]
                    context = " | ".join(row_values)
                    results.append((n, sheet_name, cell.coordinate, cell.row, context))
    wb.close()
    return results


def main() -> int:
    print(f"🔎 Buscant socis {sorted(TARGETS)} als Excels de {HISTORIC}\n")

    # First, priority files (master socis list)
    found_per_target: dict[int, list[tuple[Path, str, str, int, str]]] = {t: [] for t in TARGETS}

    for f in PRIORITY_FILES:
        if not f.exists():
            print(f"⚠️  No existeix: {f}")
            continue
        print(f"📂 {f.name}")
        for num, sheet, coord, rownum, context in scan_workbook(f, TARGETS):
            found_per_target[num].append((f, sheet, coord, rownum, context))
            print(f"   → {num} a «{sheet}!{coord}» (fila {rownum})")
            print(f"     context: {context[:200]}")

    # Second, sweep the rest of the historic folder
    print("\n🔎 Cercant també a tots els altres xls/xlsx/xlsm (pot trigar)...\n")
    other_files = [
        p for p in HISTORIC.rglob("*.xls*")
        if p.name not in {pf.name for pf in PRIORITY_FILES}
        and "$" not in p.parent.name  # skip Archive of Deleted Items
    ]
    print(f"   {len(other_files)} fitxers a processar\n")

    still_missing = {t for t in TARGETS if not found_per_target[t]}
    if still_missing:
        for f in other_files:
            if not still_missing:
                break
            try:
                for num, sheet, coord, rownum, context in scan_workbook(f, still_missing):
                    found_per_target[num].append((f, sheet, coord, rownum, context))
                    print(f"📂 {f.relative_to(HISTORIC)} → {num} a «{sheet}!{coord}»")
                    print(f"     context: {context[:200]}")
            except Exception as e:
                print(f"  ⚠️  {f.relative_to(HISTORIC)}: {e}")

    # Summary
    print("\n" + "=" * 50)
    print("RESUM")
    print("=" * 50)
    for t in sorted(TARGETS):
        hits = found_per_target[t]
        if not hits:
            print(f"❌ {t}: cap coincidència")
        else:
            print(f"✅ {t}: {len(hits)} coincidència(es)")
            # Show first 3
            for f, sheet, coord, rownum, context in hits[:3]:
                rel = f.relative_to(HISTORIC) if f.is_relative_to(HISTORIC) else f.name
                print(f"     - {rel} : {sheet}!{coord}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
