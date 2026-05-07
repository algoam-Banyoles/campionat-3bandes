"""Extreu el cens de socis del fitxer TODOS LOS DATOS BILLAR.xlsx (full DATOS SOCIOS BILLAR)
i el del 27jul22 i el genera com JSON normalitzat per fer el creuament.
"""
from pathlib import Path
from openpyxl import load_workbook
import json, re, sys, io
import warnings
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

OUT = Path(r"c:/Users/algoa/ProjectesInformatics/campionat-3bandes/tmp_census.json")

CENSUS_FILES = [
    (r"G:/Mi unidad/Foment/Historic/Foment 2022/TODOS LOS DATOS BILLAR.xlsx",
     "DATOS SOCIOS BILLAR", "todos_2022"),
    (r"G:/Mi unidad/Foment/Historic/Documentació genèrica Foment/Dades socis de la secció de billar 27jul22.xlsx",
     "Dades socis billar sense foto", "dades_27jul22"),
    (r"G:/Mi unidad/Foment/Historic/Foment 2017/Dades socis a 10-1-2017.xlsx",
     "Full1", "dades_2017"),
]

def find_header_row(ws, max_scan=15):
    """Find row with column headers (look for 'Cognoms' or 'N. Soci')."""
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx > max_scan:
            return None
        cells = [str(c).strip() if c is not None else "" for c in row]
        joined = " | ".join(cells).lower()
        if "cognoms" in joined and ("soci" in joined or "telefon" in joined or "nom" in joined):
            return r_idx, cells
    return None

def normalize_header(h):
    h = h.lower().strip()
    h = re.sub(r"[^a-z0-9]", "_", h)
    h = re.sub(r"_+", "_", h).strip("_")
    return h

def extract(path, sheet, source_id):
    print(f"\n>>> {source_id}: {Path(path).name}")
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        print(f"  Full {sheet!r} no existeix. Disponibles: {wb.sheetnames}")
        return []
    ws = wb[sheet]
    res = find_header_row(ws)
    if not res:
        print(f"  No header trobat")
        return []
    header_row, headers = res
    print(f"  Header @ r{header_row}: {headers[:14]}")
    norm_headers = [normalize_header(h) for h in headers]
    print(f"  Norm: {norm_headers[:14]}")
    rows_out = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= header_row:
            continue
        cells = list(row)
        # require at least cognoms or nom
        rec = {}
        for h, c in zip(norm_headers, cells):
            if not h:
                continue
            if c is None or str(c).strip() == "":
                continue
            rec[h] = str(c).strip() if not isinstance(c, (int, float)) else c
        if not rec:
            continue
        if not any(rec.get(k) for k in ("cognoms", "cognom", "nom", "n_soci", "num_soci")):
            continue
        rec["__source"] = source_id
        rows_out.append(rec)
    print(f"  → {len(rows_out)} files extretes")
    wb.close()
    return rows_out

all_rows = []
for path, sheet, sid in CENSUS_FILES:
    try:
        all_rows.extend(extract(path, sheet, sid))
    except Exception as e:
        print(f"ERROR {sid}: {e}")

OUT.write_text(json.dumps(all_rows, ensure_ascii=False, indent=1, default=str), encoding="utf-8")
print(f"\nTotal: {len(all_rows)} files. Escrit a {OUT.name}")

# Print first 3 of each source
for sid in {r["__source"] for r in all_rows}:
    sample = [r for r in all_rows if r["__source"] == sid][:3]
    print(f"\nMostra {sid}:")
    for r in sample:
        print(f"  {r}")
