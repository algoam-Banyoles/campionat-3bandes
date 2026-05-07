"""Escaneja els Excels de G:/Mi unidad/Foment/Historic i exporta un resum JSON
amb fulls, capçaleres i mostres de dades. Sortida a tmp_excel_report.json.

No modifica res. Només lectura.
"""
from __future__ import annotations
import json
import os
import sys
from pathlib import Path
import warnings

warnings.filterwarnings("ignore")

ROOT = Path(r"G:/Mi unidad/Foment/Historic")
OUT = Path(r"c:/Users/algoa/ProjectesInformatics/campionat-3bandes/tmp_excel_report.json")

# Skip clearly irrelevant files (memoria economica, fotos, etc.)
SKIP_PATTERNS = (
    "memoria_economica",
    "Pressupost",
    "Justificació",
    "Subvenció",
    "Plantillas",
    "Acta de partit",
    "NÚMEROS GRANDES",
    "Mòbils",
    "TELEFONOS",
    "Usuaris de taqu",
    "taqueras",
    "taquillas",
    "12 Hores",
    "Llibre.xlsx",  # generic
    "copesFoment",  # trofeus
)

def should_skip(p: Path) -> bool:
    name = p.name
    for pat in SKIP_PATTERNS:
        if pat.lower() in name.lower():
            return True
    return False


def safe_str(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() == "none":
        return None
    return s[:200]  # cap


def read_xlsx(path: Path):
    """Read .xlsx / .xlsm with openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    info = {"sheets": []}
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            rows = []
            row_count = 0
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                row_count += 1
                if r_idx < 30:
                    rows.append([safe_str(c) for c in row[:15]])
                if r_idx > 200:
                    break
            info["sheets"].append({
                "name": sheet_name,
                "rows_sampled": row_count,
                "preview": rows,
            })
        except Exception as e:
            info["sheets"].append({"name": sheet_name, "error": str(e)[:200]})
    wb.close()
    return info


def read_xls(path: Path):
    """Read legacy .xls with xlrd."""
    import xlrd
    book = xlrd.open_workbook(str(path), formatting_info=False, on_demand=True)
    info = {"sheets": []}
    for sheet_name in book.sheet_names():
        try:
            sh = book.sheet_by_name(sheet_name)
            rows = []
            for r in range(min(30, sh.nrows)):
                rows.append([safe_str(sh.cell_value(r, c)) for c in range(min(15, sh.ncols))])
            info["sheets"].append({
                "name": sheet_name,
                "rows_sampled": sh.nrows,
                "preview": rows,
            })
        except Exception as e:
            info["sheets"].append({"name": sheet_name, "error": str(e)[:200]})
    return info


def read_xlsb(path: Path):
    """Read .xlsb with pyxlsb if available."""
    try:
        from pyxlsb import open_workbook
    except ImportError:
        return {"error": "pyxlsb not installed"}
    info = {"sheets": []}
    with open_workbook(str(path)) as wb:
        for sheet_name in wb.sheets:
            try:
                with wb.get_sheet(sheet_name) as sh:
                    rows = []
                    for r_idx, row in enumerate(sh.rows()):
                        if r_idx >= 30:
                            break
                        rows.append([safe_str(c.v) for c in row[:15]])
                    info["sheets"].append({"name": sheet_name, "preview": rows})
            except Exception as e:
                info["sheets"].append({"name": sheet_name, "error": str(e)[:200]})
    return info


def main():
    files = []
    for ext in ("*.xlsx", "*.xlsm", "*.xls", "*.xlsb"):
        files.extend(ROOT.rglob(ext))

    report = {"root": str(ROOT), "files": []}
    for p in files:
        if should_skip(p):
            continue
        rel = p.relative_to(ROOT)
        entry = {"path": str(rel), "size_kb": round(p.stat().st_size / 1024, 1)}
        try:
            ext = p.suffix.lower()
            if ext in (".xlsx", ".xlsm"):
                entry.update(read_xlsx(p))
            elif ext == ".xls":
                entry.update(read_xls(p))
            elif ext == ".xlsb":
                entry.update(read_xlsb(p))
        except Exception as e:
            entry["error"] = f"{type(e).__name__}: {e}"[:300]
        report["files"].append(entry)
        print(f"OK {rel}", flush=True)

    OUT.write_text(json.dumps(report, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\nWrote {OUT} ({OUT.stat().st_size//1024} KB), {len(report['files'])} files")


if __name__ == "__main__":
    main()
