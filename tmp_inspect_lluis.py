"""Inspecciona els fitxers Lluís González per confirmar temporada/modalitat reals.
Llegeix les capçaleres reals dins els fulls (TEMPORADA, MODALITAT, CATEGORIA).
"""
from pathlib import Path
from openpyxl import load_workbook
import warnings, json
warnings.filterwarnings("ignore")

ROOT = Path(r"G:/Mi unidad/Foment/Historic/LUIS GONZALEZ Foment Excel")

files = list(ROOT.rglob("*.xlsx"))

def inspect(path: Path):
    print(f"\n{'='*80}\n{path.relative_to(ROOT)}\n{'='*80}")
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print(f"ERROR: {e}")
        return
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n  Full: '{sname}'")
        # Read first 25 rows looking for TEMPORADA/MODALITAT/CATEGORIA
        rows_with_meta = []
        data_rows = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx >= 25:
                break
            cells = [str(c).strip() if c is not None else "" for c in row[:12]]
            joined = " | ".join(cells)
            if any(k in joined.upper() for k in ("TEMPORADA", "MODALITAT", "CATEGORIA", "MODALIDAD")):
                rows_with_meta.append((r_idx, joined[:200]))
            data_rows.append(cells)
        if rows_with_meta:
            print("    Metadata trobada:")
            for r, j in rows_with_meta:
                print(f"      [r{r}] {j}")
        else:
            print("    Sense metadata estructurada. Primeres 8 files:")
            for r, cells in enumerate(data_rows[:8]):
                non_empty = " | ".join(c for c in cells if c)[:160]
                if non_empty:
                    print(f"      [r{r}] {non_empty}")
    wb.close()

for f in files:
    inspect(f)
